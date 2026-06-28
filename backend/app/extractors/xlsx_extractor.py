"""XLSX extractor: iterates all sheets, fills merged cells, detects buried headers."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from app.extractors.base import BaseExtractor, ExtractResult
from app.extractors.common.table_to_rows import table_to_rows


def _read_sheet(ws) -> list[list]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # propagate merged-cell top-left value into the merged range (read-only off here)
    try:
        for rng in ws.merged_cells.ranges:
            top = ws.cell(row=rng.min_row, column=rng.min_col).value
            if top is None:
                continue
            for rr in range(rng.min_row, rng.max_row + 1):
                for cc in range(rng.min_col, rng.max_col + 1):
                    ri, ci = rr - 1, cc - 1
                    if 0 <= ri < len(rows) and 0 <= ci < len(rows[ri]):
                        if rows[ri][ci] is None:
                            rows[ri][ci] = top
    except Exception:
        pass
    return rows


class XlsxExtractor(BaseExtractor):
    supported_extensions = {"xlsx", "xlsm"}

    def extract(self, path: Path, progress=None, should_cancel=None) -> ExtractResult:
        result = ExtractResult()
        wb = openpyxl.load_workbook(path, data_only=True)
        result.meta["sheets"] = wb.sheetnames
        for ws in wb.worksheets:
            table = _read_sheet(ws)
            rows, warns = table_to_rows(
                table, method="xlsx", source_prefix=f"sheet={ws.title};"
            )
            for row in rows:
                result.add_row(row)
            result.warnings.extend(warns)
        wb.close()
        if not result.rows:
            result.warnings.append("xlsx: no rows extracted from any sheet")
        return result
