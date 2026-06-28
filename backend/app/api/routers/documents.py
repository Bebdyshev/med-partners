from __future__ import annotations

import json
import mimetypes
import queue
import threading
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_db
from app.config import settings
from app.db.session import session_scope
from app.models import PriceDocument
from app.schemas.dto import DocumentOut
from app.services.report import compute_document_breakdown, compute_partner_breakdown, compute_report

router = APIRouter()

_MEDIA = {
    "pdf": "application/pdf",
    "scan_pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xls": "application/vnd.ms-excel",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


def _resolve_stored(stored_path: str) -> Path | None:
    """Locate the original upload. `stored_path` may be an absolute path from another
    host (DB restored from a dump) — fall back to re-rooting the `raw_files/<uuid>/<file>`
    tail under the configured storage dir so previews work on the server too."""
    p = Path(stored_path)
    if p.is_file():
        return p
    s = stored_path.replace("\\", "/")
    idx = s.rfind("raw_files/")
    if idx != -1:
        cand = settings.raw_files_dir / s[idx + len("raw_files/"):]
        if cand.is_file():
            return cand
    return None


@router.get("/documents", response_model=list[DocumentOut])
def list_documents(
    status: str | None = Query(None),
    limit: int = Query(200, le=2000),
    offset: int = 0,
    db: Session = Depends(get_db),
):
    stmt = select(PriceDocument)
    if status:
        stmt = stmt.where(PriceDocument.status == status)
    stmt = stmt.order_by(PriceDocument.created_at.desc()).limit(limit).offset(offset)
    docs = db.execute(stmt).scalars().all()
    return [_to_out(d) for d in docs]


@router.get("/documents/{doc_id}", response_model=DocumentOut)
def get_document(doc_id: uuid.UUID, db: Session = Depends(get_db)):
    doc = db.get(PriceDocument, doc_id)
    if doc is None:
        raise HTTPException(404, "document not found")
    return _to_out(doc)


@router.get("/documents/{doc_id}/file")
def get_document_file(doc_id: uuid.UUID, db: Session = Depends(get_db)):
    """Serve the immutable original upload — opens inline (PDF) or downloads (Excel/Word)."""
    doc = db.get(PriceDocument, doc_id)
    if doc is None:
        raise HTTPException(404, "document not found")
    path = _resolve_stored(doc.stored_path)
    if path is None:
        raise HTTPException(404, "stored file missing")
    fmt = doc.file_format.value if hasattr(doc.file_format, "value") else str(doc.file_format)
    media = _MEDIA.get(fmt) or mimetypes.guess_type(doc.source_filename)[0] or "application/octet-stream"
    return FileResponse(
        path,
        media_type=media,
        filename=doc.source_filename,
        content_disposition_type="inline",
    )


@router.get("/demo-file")
def demo_file():
    """Serve a small scanned price-list for the live demo (smallest PDF in data/)."""
    data_dir = Path(__file__).resolve().parents[4] / "data"
    pdfs = [p for p in data_dir.iterdir() if p.suffix.lower() == ".pdf"] if data_dir.is_dir() else []
    if not pdfs:
        raise HTTPException(404, "no demo file available")
    preferred = data_dir / "Клиника 5 прайс 2025.pdf"
    target = preferred if preferred.is_file() else min(pdfs, key=lambda p: p.stat().st_size)
    return FileResponse(
        target, media_type="application/pdf",
        filename="demo-scan.pdf", content_disposition_type="inline",
    )


@router.get("/documents/{doc_id}/page/{pageno}")
def document_page(doc_id: uuid.UUID, pageno: int, db: Session = Depends(get_db)):
    """Render one PDF page to PNG (~150 DPI), cached, for the scanner visual."""
    doc = db.get(PriceDocument, doc_id)
    if doc is None:
        raise HTTPException(404, "document not found")
    path = _resolve_stored(doc.stored_path)
    if path is None:
        raise HTTPException(404, "stored file missing")
    fmt = doc.file_format.value if hasattr(doc.file_format, "value") else str(doc.file_format)
    if "pdf" not in fmt and path.suffix.lower() != ".pdf":
        raise HTTPException(400, "page preview is only available for PDF")
    cache = settings.derived_dir / f"page_{doc_id}_{pageno}.png"
    if not cache.exists():
        from app.extractors.vision_extract import _render_png
        try:
            png = _render_png(path, pageno - 1, dpi=150)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(404, f"cannot render page {pageno}: {exc}")
        settings.derived_dir.mkdir(parents=True, exist_ok=True)
        cache.write_bytes(png)
    return FileResponse(cache, media_type="image/png")


@router.get("/documents/{doc_id}/process-stream")
def process_stream(doc_id: uuid.UUID):
    """Run the pipeline in a worker thread and stream live progress events (SSE-style).

    Pair with POST /upload?process=false: upload the file, then open this stream to
    watch extraction → OCR (per page) → normalization tallies → done in real time."""
    q: queue.Queue = queue.Queue()
    sentinel = object()

    def run() -> None:
        try:
            with session_scope() as tdb:
                doc = tdb.get(PriceDocument, doc_id)
                if doc is None:
                    q.put({"stage": "error", "message": "document not found"})
                    return
                from app.services.processing import process_document
                process_document(tdb, doc, progress=lambda e: q.put(e))
        except Exception as exc:  # noqa: BLE001 — surface to the client, never crash silently
            q.put({"stage": "error", "message": f"{type(exc).__name__}: {exc}"})
        finally:
            q.put(sentinel)

    threading.Thread(target=run, daemon=True).start()

    def gen():
        yield ": keep-alive\n\n"
        while True:
            ev = q.get()
            if ev is sentinel:
                break
            yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )


def _ref_kv(ref: str) -> dict:
    kv: dict = {}
    for part in (ref or "").split(";"):
        if "=" in part:
            k, v = part.split("=", 1)
            kv[k] = v
    return kv


def _trim(v) -> str:
    s = "" if v is None else str(v).strip()
    return s[:80]


def _finish(rows: list[dict], label: str, target: int) -> dict:
    """Drop trailing all-empty columns so the fragment isn't needlessly wide."""
    width = 0
    for r in rows:
        last = 0
        for i, c in enumerate(r["cells"]):
            if c:
                last = i + 1
        width = max(width, last)
    for r in rows:
        r["cells"] = r["cells"][:width]
    return {"kind": "table", "label": label, "target": target, "rows": rows}


@router.get("/documents/{doc_id}/preview")
def document_preview(doc_id: uuid.UUID, ref: str = Query("", description="source_ref of the item"),
                     db: Session = Depends(get_db)):
    """A focused fragment of an Excel/Word source around the item's row — so the
    operator sees the line in context without a native Office viewer (ТЗ 4.4)."""
    doc = db.get(PriceDocument, doc_id)
    if doc is None:
        raise HTTPException(404, "document not found")
    path = _resolve_stored(doc.stored_path)
    if path is None:
        raise HTTPException(404, "stored file missing")
    fmt = doc.file_format.value if hasattr(doc.file_format, "value") else str(doc.file_format)
    kv = _ref_kv(ref)
    row = int(kv.get("row", 0) or 0)
    W, MAXC = 4, 8  # ±4 rows of context, up to 8 columns

    try:
        suffix = path.suffix.lower()
        if fmt == "xlsx" or suffix == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = kv.get("sheet")
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            lo = max(1, row - W)
            rows = [{"n": lo + i, "cells": [_trim(v) for v in vals]}
                    for i, vals in enumerate(ws.iter_rows(min_row=lo, max_row=row + W, max_col=MAXC, values_only=True))]
            wb.close()
            return _finish(rows, f"лист «{ws.title}»", row)

        if fmt == "xls" or suffix == ".xls":
            import xlrd
            book = xlrd.open_workbook(str(path))
            sheet = kv.get("sheet")
            sh = book.sheet_by_name(sheet) if sheet and sheet in book.sheet_names() else book.sheet_by_index(0)
            lo, hi = max(1, row - W), min(sh.nrows, row + W)
            rows = [{"n": r, "cells": [_trim(v) for v in sh.row_values(r - 1)[:MAXC]]} for r in range(lo, hi + 1)]
            return _finish(rows, f"лист «{sheet or sh.name}»", row)

        if fmt in ("docx",) or suffix == ".docx":
            import zipfile
            from app.extractors.docx_extractor import _iter_tables
            with zipfile.ZipFile(path) as z:
                xml = z.read("word/document.xml")
            ti = int(kv.get("table", 0) or 0)
            tables = list(_iter_tables(xml))
            if ti >= len(tables):
                return {"kind": "unsupported"}
            tbl = tables[ti]
            lo, hi = max(1, row - W), min(len(tbl), row + W)
            rows = [{"n": r, "cells": [_trim(c) for c in (tbl[r - 1] or [])[:MAXC]]} for r in range(lo, hi + 1)]
            return _finish(rows, f"таблица {ti + 1}", row)
    except Exception as e:  # noqa: BLE001 — preview is best-effort
        return {"kind": "unsupported", "error": str(e)}
    return {"kind": "unsupported"}


@router.get("/dashboard/stats")
def dashboard_stats(db: Session = Depends(get_db)):
    return compute_report()


@router.get("/dashboard/documents")
def dashboard_documents():
    """Per-document composition + provenance + category mix for the dashboard ledger."""
    return compute_document_breakdown()


@router.get("/dashboard/partners")
def dashboard_partners():
    """Per-partner rollup (positions, auto-match rate, price freshness) for the directory."""
    return compute_partner_breakdown()


def _to_out(d: PriceDocument) -> DocumentOut:
    return DocumentOut(
        id=d.id,
        partner_id=d.partner_id,
        source_filename=d.source_filename,
        file_format=d.file_format.value,
        status=d.status.value,
        year=d.year,
        parsed_at=d.parsed_at,
        method_summary=d.method_summary or {},
        warnings=d.warnings or [],
    )
